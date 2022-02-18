import pandas as pd
import sympy
from datetime import datetime
import calendar
import openpyxl
from openpyxl.writer.excel import save_workbook


filename = 'task_support.xlsx'


def save_answers(answers: tuple) -> None:
    wb = openpyxl.load_workbook(filename)
    ws = wb.create_sheet('Answers')
    headers = ['even_count', 'count_prime', 'num_less_5', 'count_tue', 'count_tue_2', 'count_last_tue']
    for i, header in enumerate(headers):
        ws.cell(1, i + 1, header)
        ws.cell(2, i + 1, str(answers[i]))
    save_workbook(wb, filename)


def is_last_tue(date):
    last_tue = max(week[1] for week in calendar.monthcalendar(date.year, date.month))
    if date.day == last_tue:
        return True
    return False


def calculate():
    df = pd.read_excel(filename, sheet_name='Tasks')

    num_1 = pd.to_numeric(df['num1'][1:])
    even_count = (num_1 % 2 == 0).sum()

    num_2 = df['num2'][1:].astype(int).to_list()
    simple_count = sum((1 for i in num_2 if sympy.isprime(i)))

    num_3 = df['num3'][1:].replace(r'\s+', '', regex=True)
    num_3 = num_3.str.replace(',', '.').astype(float)
    num_less_5 = (num_3 < .5).sum()

    date_1 = df['date1'][1:].astype(str).to_list()
    count_tue = sum((1 for i in date_1 if datetime.strptime(i, '%a %b %d %H:%M:%S %Y').day == 2))

    date_2 = df['date2'][1:].astype(str).to_list()
    count_tue_2 = sum((1 for i in date_2 if datetime.strptime(i, '%Y-%m-%d %H:%M:%S.%f').day == 2))

    date_3 = df['date3'][1:].astype(str).to_list()
    count_last_tue = sum((1 for i in date_3 if is_last_tue(datetime.strptime(i, '%m-%d-%Y'))))

    return even_count, simple_count, num_less_5, count_tue, count_tue_2, count_last_tue


if __name__ == '__main__':
    save_answers(calculate())
